#!/usr/bin/env python3
"""
VectorCAST Report Analyser

A comprehensive tool for analyzing VectorCAST test reports and generating 
detailed Excel summaries with directory tree structures and report metrics.

This tool scans directories for VectorCAST report files, extracts relevant 
information, and creates structured Excel reports for test analysis and 
project management.

Author: Suduli Kumar
Version: 2.0.0
License: MIT
"""

import os
import re
import sys
import logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple, Any
from datetime import datetime
import argparse

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError as e:
    print(f"Required package missing: {e}")
    print("Install required packages: pip install pandas openpyxl")
    sys.exit(1)


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('vectorcast_analyzer.log'),
        logging.StreamHandler(sys.stdout)
    ]
)

logger = logging.getLogger(__name__)


class VectorCASTReportAnalyser:
    """
    VectorCAST Report Analyser for extracting and organizing test report data.
    
    This class provides comprehensive functionality for scanning directories,
    extracting VectorCAST report information, and generating structured Excel reports.
    """
    
    def __init__(self, root_directory: str = "."):
        """
        Initialize the VectorCAST Report Analyser.
        
        Args:
            root_directory (str): Root directory to scan for reports
        """
        self.root_directory = Path(root_directory).resolve()
        self.report_patterns = {
            'full_report': r'.*\.Full_Report\.html$',
            'metrics_report': r'.*\.Metrics_Report\.html$',
            'testcase_report': r'.*\.Testcase_Management_Report\.html$',
            'coverage_report': r'.*\.Coverage_Report\.html$',
            'execution_report': r'.*\.Execution_Report\.html$'
        }
        self.extracted_data: Dict[str, List[str]] = {}
        self.directory_tree: List[str] = []
    
    def generate_directory_tree(self, output_file: Optional[str] = None) -> List[str]:
        """
        Generate a directory tree structure of the specified directory.
        
        Args:
            output_file (str, optional): File to save the tree structure
            
        Returns:
            List[str]: Directory tree structure as list of strings
        """
        logger.info(f"Generating directory tree for: {self.root_directory}")
        
        tree_lines = []
        tree_lines.append(f"Directory Tree for: {self.root_directory}")
        tree_lines.append("=" * 50)
        
        def add_tree_line(path: Path, prefix: str = "") -> None:
            """Add a line to the tree structure."""
            if path.is_dir():
                tree_lines.append(f"{prefix}📁 {path.name}/")
                try:
                    items = sorted(path.iterdir(), key=lambda x: (x.is_file(), x.name.lower()))
                    for i, item in enumerate(items):
                        is_last = i == len(items) - 1
                        new_prefix = prefix + ("    " if is_last else "│   ")
                        connector = "└── " if is_last else "├── "
                        
                        if item.is_dir():
                            tree_lines.append(f"{prefix}{connector}📁 {item.name}/")
                            add_tree_line(item, new_prefix)
                        else:
                            icon = self._get_file_icon(item.suffix)
                            tree_lines.append(f"{prefix}{connector}{icon} {item.name}")
                            
                except PermissionError:
                    tree_lines.append(f"{prefix}    [Permission Denied]")
        
        add_tree_line(self.root_directory)
        
        if output_file:
            try:
                with open(output_file, 'w', encoding='utf-8') as f:
                    f.write('\n'.join(tree_lines))
                logger.info(f"Directory tree saved to: {output_file}")
            except Exception as e:
                logger.error(f"Failed to save directory tree: {e}")
        
        self.directory_tree = tree_lines
        return tree_lines
    
    def _get_file_icon(self, extension: str) -> str:
        """Get appropriate icon for file extension."""
        icons = {
            '.html': '🌐',
            '.xml': '📄',
            '.json': '📋',
            '.csv': '📊',
            '.xlsx': '📈',
            '.py': '🐍',
            '.txt': '📝',
            '.log': '📃'
        }
        return icons.get(extension.lower(), '📄')
    
    def extract_file_names(self, content: str, pattern: str) -> List[str]:
        """
        Extract filenames from content based on a regex pattern.
        
        Args:
            content (str): Content to search in
            pattern (str): Regex pattern to match
            
        Returns:
            List[str]: List of matching filenames
        """
        try:
            matches = re.findall(pattern, content, re.MULTILINE | re.IGNORECASE)
            return list(set(matches))  # Remove duplicates
        except re.error as e:
            logger.error(f"Invalid regex pattern '{pattern}': {e}")
            return []
    
    def scan_directory_for_reports(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        Scan directory for VectorCAST report files and extract metadata.
        
        Returns:
            Dict[str, List[Dict[str, Any]]]: Dictionary containing report information
        """
        logger.info("Scanning directory for VectorCAST reports...")
        
        report_data = {
            'full_reports': [],
            'metrics_reports': [],
            'testcase_reports': [],
            'coverage_reports': [],
            'execution_reports': []
        }
        
        try:
            for root, dirs, files in os.walk(self.root_directory):
                root_path = Path(root)
                
                for file in files:
                    file_path = root_path / file
                    
                    # Check against each pattern
                    for report_type, pattern in self.report_patterns.items():
                        if re.match(pattern, file, re.IGNORECASE):
                            report_info = self._extract_report_metadata(file_path)
                            
                            # Map report types to data keys
                            data_key = f"{report_type.replace('_report', '')}_reports"
                            if report_type == 'testcase_report':
                                data_key = 'testcase_reports'
                            
                            report_data[data_key].append(report_info)
                            break
        
        except Exception as e:
            logger.error(f"Error scanning directory: {e}")
        
        # Log summary
        total_reports = sum(len(reports) for reports in report_data.values())
        logger.info(f"Found {total_reports} VectorCAST reports")
        
        for report_type, reports in report_data.items():
            if reports:
                logger.info(f"  - {report_type}: {len(reports)} files")
        
        return report_data
    
    def _extract_report_metadata(self, file_path: Path) -> Dict[str, Any]:
        """
        Extract metadata from a report file.
        
        Args:
            file_path (Path): Path to the report file
            
        Returns:
            Dict[str, Any]: Report metadata
        """
        try:
            stat_info = file_path.stat()
            return {
                'filename': file_path.name,
                'full_path': str(file_path),
                'relative_path': str(file_path.relative_to(self.root_directory)),
                'size_bytes': stat_info.st_size,
                'size_mb': round(stat_info.st_size / (1024 * 1024), 2),
                'modified_date': datetime.fromtimestamp(stat_info.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                'directory': str(file_path.parent.relative_to(self.root_directory))
            }
        except Exception as e:
            logger.warning(f"Could not extract metadata for {file_path}: {e}")
            return {
                'filename': file_path.name,
                'full_path': str(file_path),
                'relative_path': str(file_path.relative_to(self.root_directory)),
                'size_bytes': 0,
                'size_mb': 0,
                'modified_date': 'Unknown',
                'directory': str(file_path.parent.relative_to(self.root_directory))
            }
    
    def pad_lists_to_same_length(self, *lists: List[Any]) -> List[List[Any]]:
        """
        Pad lists with None values until they all have the same length.
        
        Args:
            *lists: Variable number of lists to pad
            
        Returns:
            List[List[Any]]: Padded lists
        """
        if not lists:
            return []
        
        max_length = max(len(lst) for lst in lists)
        padded_lists = []
        
        for lst in lists:
            padded_list = lst + [None] * (max_length - len(lst))
            padded_lists.append(padded_list)
        
        return padded_lists
    
    def create_excel_report(self, output_filename: str = "vectorcast_analysis.xlsx") -> None:
        """
        Create a comprehensive Excel report with multiple sheets.
        
        Args:
            output_filename (str): Output Excel filename
        """
        logger.info(f"Creating Excel report: {output_filename}")
        
        try:
            # Generate directory tree and scan for reports
            self.generate_directory_tree()
            report_data = self.scan_directory_for_reports()
            
            # Create workbook
            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                
                # Create summary sheet
                self._create_summary_sheet(writer, report_data)
                
                # Create detailed sheets for each report type
                for report_type, reports in report_data.items():
                    if reports:
                        self._create_report_sheet(writer, report_type, reports)
                
                # Create directory tree sheet
                self._create_directory_tree_sheet(writer)
            
            # Apply formatting
            self._format_excel_file(output_filename)
            
            logger.info(f"Excel report created successfully: {output_filename}")
            
        except Exception as e:
            logger.error(f"Failed to create Excel report: {e}")
            raise
    
    def _create_summary_sheet(self, writer: pd.ExcelWriter, report_data: Dict[str, List[Dict[str, Any]]]) -> None:
        """Create summary sheet with overview statistics."""
        summary_data = {
            'Report Type': [],
            'Count': [],
            'Total Size (MB)': [],
            'Average Size (MB)': [],
            'Latest Modified': []
        }
        
        for report_type, reports in report_data.items():
            if reports:
                total_size = sum(report['size_mb'] for report in reports)
                avg_size = total_size / len(reports) if reports else 0
                latest_modified = max(report['modified_date'] for report in reports)
                
                summary_data['Report Type'].append(report_type.replace('_', ' ').title())
                summary_data['Count'].append(len(reports))
                summary_data['Total Size (MB)'].append(round(total_size, 2))
                summary_data['Average Size (MB)'].append(round(avg_size, 2))
                summary_data['Latest Modified'].append(latest_modified)
        
        # Add total row
        total_count = sum(len(reports) for reports in report_data.values())
        total_size = sum(sum(report['size_mb'] for report in reports) for reports in report_data.values())
        
        summary_data['Report Type'].append('TOTAL')
        summary_data['Count'].append(total_count)
        summary_data['Total Size (MB)'].append(round(total_size, 2))
        summary_data['Average Size (MB)'].append(round(total_size / total_count if total_count > 0 else 0, 2))
        summary_data['Latest Modified'].append('')
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_report_sheet(self, writer: pd.ExcelWriter, report_type: str, reports: List[Dict[str, Any]]) -> None:
        """Create sheet for specific report type."""
        df_reports = pd.DataFrame(reports)
        sheet_name = report_type.replace('_', ' ').title()[:31]  # Excel sheet name limit
        df_reports.to_excel(writer, sheet_name=sheet_name, index=False)
    
    def _create_directory_tree_sheet(self, writer: pd.ExcelWriter) -> None:
        """Create directory tree sheet."""
        tree_df = pd.DataFrame({'Directory Structure': self.directory_tree})
        tree_df.to_excel(writer, sheet_name='Directory Tree', index=False)
    
    def _format_excel_file(self, filename: str) -> None:
        """Apply formatting to the Excel file."""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filename)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Format each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format headers
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            
        except Exception as e:
            logger.warning(f"Could not apply Excel formatting: {e}")


def main():
    """Main entry point of the application."""
    parser = argparse.ArgumentParser(description='VectorCAST Report Analyser')
    parser.add_argument(
        '-d', '--directory',
        default='.',
        help='Root directory to scan (default: current directory)'
    )
    parser.add_argument(
        '-o', '--output',
        default='vectorcast_analysis.xlsx',
        help='Output Excel filename (default: vectorcast_analysis.xlsx)'
    )
    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='Enable verbose logging'
    )
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    try:
        logger.info("Starting VectorCAST Report Analysis")
        logger.info(f"Scanning directory: {args.directory}")
        
        analyzer = VectorCASTReportAnalyser(args.directory)
        analyzer.create_excel_report(args.output)
        
        logger.info("Analysis completed successfully!")
        
    except KeyboardInterrupt:
        logger.info("Analysis interrupted by user")
        sys.exit(0)
    except Exception as e:
        logger.error(f"Analysis failed: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
